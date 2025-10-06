import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('System Chart.xlsx')

# Get all sheet names
print("=== WORKBOOK SHEETS ===")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print()

# Dictionary to store all data
all_data = {}

# Iterate through each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")

    ws = wb[sheet_name]
    sheet_data = {
        'name': sheet_name,
        'max_row': ws.max_row,
        'max_column': ws.max_column,
        'cells': [],
        'merged_cells': []
    }

    # Get merged cell ranges
    for merged_range in ws.merged_cells.ranges:
        sheet_data['merged_cells'].append(str(merged_range))
        print(f"Merged cells: {merged_range}")

    print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")
    print("\nCell Contents:")

    # Read all non-empty cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell_info = {
                    'coordinate': cell.coordinate,
                    'row': cell.row,
                    'column': cell.column,
                    'column_letter': cell.column_letter,
                    'value': str(cell.value),
                    'fill': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color else None,
                    'font_color': str(cell.font.color.rgb) if cell.font and cell.font.color else None
                }
                sheet_data['cells'].append(cell_info)
                print(f"  {cell.coordinate}: {cell.value}")

    # Print the grid view
    print(f"\nGrid View (Sheet: {sheet_name}):")
    print("-" * 80)

    # Create a grid representation
    grid = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                if cell.row not in grid:
                    grid[cell.row] = {}
                grid[cell.row][cell.column] = str(cell.value)

    # Print grid
    for row_num in sorted(grid.keys()):
        row_data = grid[row_num]
        max_col = max(row_data.keys())
        row_str = ""
        for col in range(1, max_col + 1):
            cell_val = row_data.get(col, "")
            row_str += f"{cell_val:12s} "
        print(f"Row {row_num:2d}: {row_str}")

    all_data[sheet_name] = sheet_data

# Save to JSON file for easy access
with open('system_chart_data.json', 'w') as f:
    json.dump(all_data, f, indent=2)

print("\n" + "="*60)
print("Data exported to system_chart_data.json")
print("="*60)
